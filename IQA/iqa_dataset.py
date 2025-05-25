import csv
import os
import pandas as pd
import numpy as np
from scipy import io
import torch.utils.data as data
from PIL import Image
import torch
import torchvision.transforms.functional as F
from openpyxl import load_workbook


class KONIQDATASET(data.Dataset):
    def __init__(self, root, index, patch_num, transform=None):
        super(KONIQDATASET, self).__init__()

        self.data_path = root
        imgname = []
        mos_all = []
        csv_file = os.path.join(root, "koniq10k_scores_and_distributions.csv")
        with open(csv_file) as f:
            reader = csv.DictReader(f)
            for row in reader:
                imgname.append(row["image_name"])
                mos = np.array(float(row["MOS_zscore"])).astype(np.float32)
                mos_all.append(mos)

        sample = []
        for _, item in enumerate(index):
            for _ in range(patch_num):
                sample.append(
                    (os.path.join(root, "1024x768", imgname[item]), mos_all[item])
                )

        self.samples = sample
        self.transform = transform

    def _load_image(self, path):
        try:
            im = Image.open(path).convert("RGB")
        except:
            print("ERROR IMG LOADED: ", path)
            random_img = np.random.rand(224, 224, 3) * 255
            im = Image.fromarray(np.uint8(random_img))
        return im

    def __getitem__(self, index):
        """
        Args:
            index (int): Index

        Returns:
            tuple: (sample, target) where target is class_index of the target class.
        """
        path, target = self.samples[index]
        sample = self._load_image(path)
        
        sample = self.transform[0](sample)
        sample = self.transform[1](sample)
        sample = self.transform[2](sample)
        return sample, target

    def __len__(self):
        length = len(self.samples)
        return length


class LIVECDATASET(data.Dataset):
    def __init__(self, root, index, patch_num, transform=None, score_k=2):

        imgpath = io.loadmat(os.path.join(root, "Data", "AllImages_release.mat"))
        imgpath = imgpath["AllImages_release"]
        imgpath = imgpath[7:1169]
        mos = io.loadmat(os.path.join(root, "Data", "AllMOS_release.mat"))
        labels = mos["AllMOS_release"].astype(np.float32)
        labels = labels[0][7:1169]

        sample = []
        for i, item in enumerate(index):
            for aug in range(patch_num):
                sample.append(
                    (os.path.join(root, "Images", imgpath[item][0][0]), labels[item])
                )

        self.score_k = score_k
        self.bins = np.linspace(0, 100, self.score_k + 1)

        self.samples = sample
        self.transform = transform

    def _load_image(self, path, mode="RGB"):
        try:
            im = Image.open(path).convert(mode)
        except:
            print("ERROR IMG LOADED: ", path)
            random_img = np.random.rand(224, 224, 3) * 255
            im = Image.fromarray(np.uint8(random_img))
        return im

    def _depth_path_converter(self, img_path):
        directory, filename = os.path.split(img_path)
        name, ext = os.path.splitext(filename)

        depth_dir = os.path.join(os.path.dirname(directory), "Depths")
        depths_file = f"{name}-depth.png"
        orders_file = f"{name}-onehot_depth.npy"

        return os.path.join(depth_dir, depths_file), os.path.join(
            depth_dir, orders_file
        )

    def score_to_onehot(self, target_score):
        index = np.digitize([target_score], self.bins) - 1
        index = min(index[0], self.score_k - 1)

        one_hot = np.zeros(self.score_k, dtype=np.float32)
        one_hot[index] = 1.0
        one_hot = torch.from_numpy(one_hot).float()
        return one_hot

    def __getitem__(self, index):
        """
        Args:
            index (int): Index
        Returns:
            tuple: (sample, target) where target is class_index of the target class.
        """
        path, target = self.samples[index]
        sample = np.array(self._load_image(path)).astype(np.float32)
        cor_depth, cor_orders = (
            np.array(
                self._load_image(self._depth_path_converter(path)[0], mode="L")
            ).astype(np.float32),
            np.load(self._depth_path_converter(path)[1], allow_pickle=True).astype(
                np.float32
            ),
        )
        sample = torch.from_numpy(
            np.transpose(
                np.concatenate([sample, cor_depth[:, :, None], cor_orders], axis=-1),
                (2, 0, 1),
            )
        ).contiguous()
        sample = self.transform[0](sample)
        images, depths, orders = (
            (sample[:3, :, :] / 255.0).contiguous(),
            sample[3, :, :].contiguous(),
            sample[4:, :, :].contiguous(),
        )
        images = self.transform[2](images)
        return (
            images,
            # process_inverse_depth(depths),
            # orders,
            target
            # self.score_to_onehot(target),
        )

    def __len__(self):
        length = len(self.samples)
        return length


class UWIQADATASET(data.Dataset):
    def __init__(self, root, index, patch_num, transform=None):

        imgpath = io.loadmat(os.path.join(root, "Data", "AllImages_release.mat"))
        imgpath = imgpath["AllImages_release"]
        imgpath = imgpath[0:890]
        mos = io.loadmat(os.path.join(root, "Data", "AllMOS_release.mat"))
        labels = mos["AllMOS_release"].astype(np.float32)
        labels = labels[0][0:890]

        sample = []
        for i, item in enumerate(index):
            for aug in range(patch_num):
                sample.append(
                    (os.path.join(root, "Images", imgpath[item][0][0]), labels[item])
                )

        self.samples = sample
        self.transform = transform

    def _load_image(self, path):
        try:
            im = Image.open(path).convert("RGB")
        except:
            print("ERROR IMG LOADED: ", path)
            random_img = np.random.rand(224, 224, 3) * 255
            im = Image.fromarray(np.uint8(random_img))
        return im

    def __getitem__(self, index):
        """
        Args:
            index (int): Index
        Returns:
            tuple: (sample, target) where target is class_index of the target class.
        """
        path, target = self.samples[index]
        sample = self._load_image(path)
        sample = self.transform(sample)
        return sample, target

    def __len__(self):
        length = len(self.samples)
        return length


class LIVEDataset(data.Dataset):
    def __init__(self, root, index, patch_num, transform=None, score_k=2):

        refpath = os.path.join(root, "refimgs")
        refname = getFileName(refpath, ".bmp")

        jp2kroot = os.path.join(root, "jp2k")
        jp2kname = self.getDistortionTypeFileName(jp2kroot, 227)

        jpegroot = os.path.join(root, "jpeg")
        jpegname = self.getDistortionTypeFileName(jpegroot, 233)

        wnroot = os.path.join(root, "wn")
        wnname = self.getDistortionTypeFileName(wnroot, 174)

        gblurroot = os.path.join(root, "gblur")
        gblurname = self.getDistortionTypeFileName(gblurroot, 174)

        fastfadingroot = os.path.join(root, "fastfading")
        fastfadingname = self.getDistortionTypeFileName(fastfadingroot, 174)

        imgpath = jp2kname + jpegname + wnname + gblurname + fastfadingname

        dmos = io.loadmat(os.path.join(root, "dmos_realigned.mat"))
        labels = dmos["dmos_new"].astype(np.float32)

        orgs = dmos["orgs"]
        refnames_all = io.loadmat(os.path.join(root, "refnames_all.mat"))
        refnames_all = refnames_all["refnames_all"]

        refname.sort()
        sample = []

        for i in range(0, len(index)):
            train_sel = refname[index[i]] == refnames_all
            train_sel = train_sel * ~orgs.astype(np.bool_)
            train_sel = np.where(train_sel == True)
            train_sel = train_sel[1].tolist()
            for j, item in enumerate(train_sel):
                for aug in range(patch_num):
                    sample.append((imgpath[item], labels[0][item]))

        self.score_k = score_k
        self.bins = np.linspace(0, 100, self.score_k + 1)

        self.samples = sample
        self.transform = transform

    def _load_image(self, path, mode="RGB"):
        try:
            im = Image.open(path).convert(mode)
        except:
            print("ERROR IMG LOADED: ", path)
            random_img = np.random.rand(224, 224, 3) * 255
            im = Image.fromarray(np.uint8(random_img))
        return im

    def _depth_path_converter(self, img_path):
        directory, filename = os.path.split(img_path)
        name, ext = os.path.splitext(filename)

        depth_dir = directory.replace("databaserelease2", "LIVE_Depths")
        depths_file = f"{name}-depth.png"
        orders_file = f"{name}-onehot_depth.npy"

        return os.path.join(depth_dir, depths_file), os.path.join(
            depth_dir, orders_file
        )

    def score_to_onehot(self, target_score):
        index = np.digitize([target_score], self.bins) - 1
        index = min(index[0], self.score_k - 1)

        one_hot = np.zeros(self.score_k, dtype=np.float32)
        one_hot[index] = 1.0
        one_hot = torch.from_numpy(one_hot).float()
        return one_hot

    def __getitem__(self, index):
        """
        Args:
            index (int): Index
        Returns:
            tuple: (sample, target) where target is class_index of the target class.
        """
        path, target = self.samples[index]
        sample = np.array(self._load_image(path)).astype(np.float32)
        cor_depth, cor_orders = (
            np.array(
                self._load_image(self._depth_path_converter(path)[0], mode="L")
            ).astype(np.float32),
            np.load(self._depth_path_converter(path)[1], allow_pickle=True).astype(
                np.float32
            ),
        )
        sample = torch.from_numpy(
            np.transpose(
                np.concatenate([sample, cor_depth[:, :, None], cor_orders], axis=-1),
                (2, 0, 1),
            )
        ).contiguous()
        sample = self.transform[0](sample)
        images, depths, orders = (
            (sample[:3, :, :] / 255.0).contiguous(),
            sample[3, :, :].contiguous(),
            sample[4:, :, :].contiguous(),
        )
        images = self.transform[2](images)
        return (
            images,
            process_inverse_depth(depths),
            orders,
            target,
            self.score_to_onehot(target),
        )

    def __len__(self):
        length = len(self.samples)
        return length

    def getDistortionTypeFileName(self, path, num):
        filename = []
        index = 1
        for i in range(0, num):
            name = "%s%s%s" % ("img", str(index), ".bmp")
            filename.append(os.path.join(path, name))
            index = index + 1
        return filename


def getFileName(path, suffix):
    filename = []
    f_list = os.listdir(path)
    for i in f_list:
        if os.path.splitext(i)[1] == suffix:
            filename.append(i)
    return filename


class TID2013Dataset(data.Dataset):
    def __init__(self, root, index, patch_num, transform=None, score_k=2):
        refpath = os.path.join(root, "reference_images")
        refname = getTIDFileName(refpath, ".bmp.BMP")
        txtpath = os.path.join(root, "mos_with_names.txt")
        fh = open(txtpath, "r")
        imgnames = []
        target = []
        refnames_all = []
        for line in fh:
            line = line.split("\n")
            words = line[0].split()
            imgnames.append((words[1]))
            target.append(words[0])
            ref_temp = words[1].split("_")
            refnames_all.append(ref_temp[0][1:])
        labels = np.float32(target)
        refnames_all = np.array(refnames_all)

        refname.sort()
        sample = []
        for i, item in enumerate(index):
            train_sel = refname[index[i]] == refnames_all
            train_sel = np.where(train_sel == True)
            train_sel = train_sel[0].tolist()
            for j, item in enumerate(train_sel):
                for aug in range(patch_num):
                    sample.append(
                        (
                            os.path.join(root, "distorted_images", imgnames[item]),
                            labels[item],
                        )
                    )

        self.score_k = score_k
        self.bins = np.linspace(0, 100, self.score_k + 1)
        self.samples = sample
        self.transform = transform

    def _load_image(self, path, mode="RGB"):
        try:
            im = Image.open(path).convert(mode)
        except:
            print("ERROR IMG LOADED: ", path)
            random_img = np.random.rand(224, 224, 3) * 255
            im = Image.fromarray(np.uint8(random_img))
        return im

    def _depth_path_converter(self, img_path):
        directory, filename = os.path.split(img_path)
        name, ext = os.path.splitext(filename)

        depth_dir = directory.replace("distorted_images", "distorted_images_depths")
        depths_file = f"{name}-depth.png"
        orders_file = f"{name}-onehot_depth.npy"

        return os.path.join(depth_dir, depths_file), os.path.join(
            depth_dir, orders_file
        )

    def score_to_onehot(self, target_score):
        index = np.digitize([target_score], self.bins) - 1
        index = min(index[0], self.score_k - 1)

        one_hot = np.zeros(self.score_k, dtype=np.float32)
        one_hot[index] = 1.0
        one_hot = torch.from_numpy(one_hot).float()
        return one_hot

    def __getitem__(self, index):
        """
        Args:
            index (int): Index
        Returns:
            tuple: (sample, target) where target is class_index of the target class.
        """
        path, target = self.samples[index]
        sample = np.array(self._load_image(path)).astype(np.float32)
        cor_depth, cor_orders = (
            np.array(
                self._load_image(self._depth_path_converter(path)[0], mode="L")
            ).astype(np.float32),
            np.load(self._depth_path_converter(path)[1], allow_pickle=True).astype(
                np.float32
            ),
        )
        sample = torch.from_numpy(
            np.transpose(
                np.concatenate([sample, cor_depth[:, :, None], cor_orders], axis=-1),
                (2, 0, 1),
            )
        ).contiguous()
        sample = self.transform[0](sample)
        images, depths, orders = (
            (sample[:3, :, :] / 255.0).contiguous(),
            sample[3, :, :].contiguous(),
            sample[4:, :, :].contiguous(),
        )
        images = self.transform[2](images)
        return (
            images,
            process_inverse_depth(depths),
            orders,
            target,
            self.score_to_onehot(target),
        )

    def __len__(self):
        length = len(self.samples)
        return length


def getTIDFileName(path, suffix):
    filename = []
    f_list = os.listdir(path)
    for i in f_list:
        if suffix.find(os.path.splitext(i)[1]) != -1:
            filename.append(i[1:3])
    return filename


class CSIQDataset(data.Dataset):
    def __init__(self, root, index, patch_num, transform=None, score_k=2):

        refpath = os.path.join(root, "src_imgs")
        refname = getFileName(refpath, ".png")
        txtpath = os.path.join(root, "csiq_label.txt")
        fh = open(txtpath, "r")
        imgnames = []
        target = []
        refnames_all = []
        for line in fh:
            line = line.split("\n")
            words = line[0].split()
            imgnames.append((words[0]))
            target.append(words[1])
            ref_temp = words[0].split(".")
            refnames_all.append(ref_temp[0] + "." + ref_temp[-1])

        labels = np.float32(target)
        refnames_all = np.array(refnames_all)

        sample = []

        for i, item in enumerate(index):
            train_sel = refname[index[i]] == refnames_all
            train_sel = np.where(train_sel == True)
            train_sel = train_sel[0].tolist()
            for j, item in enumerate(train_sel):
                for aug in range(patch_num):
                    sample.append(
                        (
                            os.path.join(root, "dst_imgs", imgnames[item]),
                            labels[item],
                        )
                    )

        self.score_k = score_k
        self.bins = np.linspace(0, 100, self.score_k + 1)

        self.samples = sample
        self.transform = transform

    def _load_image(self, path, mode="RGB"):
        try:
            im = Image.open(path).convert(mode)
        except:
            print("ERROR IMG LOADED: ", path)
            random_img = np.random.rand(224, 224, 3) * 255
            im = Image.fromarray(np.uint8(random_img))
        return im

    def _depth_path_converter(self, img_path):
        directory, filename = os.path.split(img_path)
        name, ext = os.path.splitext(filename)

        depth_dir = directory.replace("dst_imgs", "dst_imgs_depths")
        depths_file = f"{name}-depth.png"
        orders_file = f"{name}-onehot_depth.npy"

        return os.path.join(depth_dir, depths_file), os.path.join(
            depth_dir, orders_file
        )

    def score_to_onehot(self, target_score):
        index = np.digitize([target_score], self.bins) - 1
        index = min(index[0], self.score_k - 1)

        one_hot = np.zeros(self.score_k, dtype=np.float32)
        one_hot[index] = 1.0
        one_hot = torch.from_numpy(one_hot).float()
        return one_hot

    def __getitem__(self, index):
        """
        Args:
            index (int): Index
        Returns:
            tuple: (sample, target) where target is class_index of the target class.
        """
        path, target = self.samples[index]
        sample = np.array(self._load_image(path)).astype(np.float32)
        cor_depth, cor_orders = (
            np.array(
                self._load_image(self._depth_path_converter(path)[0], mode="L")
            ).astype(np.float32),
            np.load(self._depth_path_converter(path)[1], allow_pickle=True).astype(
                np.float32
            ),
        )
        sample = torch.from_numpy(
            np.transpose(
                np.concatenate([sample, cor_depth[:, :, None], cor_orders], axis=-1),
                (2, 0, 1),
            )
        ).contiguous()
        sample = self.transform[0](sample)
        images, depths, orders = (
            (sample[:3, :, :] / 255.0).contiguous(),
            sample[3, :, :].contiguous(),
            sample[4:, :, :].contiguous(),
        )
        images = self.transform[2](images)
        return (
            images,
            process_inverse_depth(depths),
            orders,
            target,
            self.score_to_onehot(target),
        )

    def __len__(self):
        length = len(self.samples)
        return length


class KADIDDataset(data.Dataset):
    def __init__(self, root, index, patch_num, transform=None, score_k=2):
        refpath = os.path.join(root, "reference_images")
        refname = getTIDFileName(refpath, ".png.PNG")

        imgnames = []
        target = []
        refnames_all = []

        csv_file = os.path.join(root, "dmos.csv")
        with open(csv_file) as f:
            reader = csv.DictReader(f)
            for row in reader:
                imgnames.append(row["dist_img"])
                refnames_all.append(row["ref_img"][1:3])

                mos = np.float32(float(row["dmos"]))
                target.append(mos)

        # labels = np.array(target).astype(np.float32)
        refnames_all = np.array(refnames_all)

        refname.sort()
        sample = []
        for i, item in enumerate(index):
            train_sel = refname[index[i]] == refnames_all
            train_sel = np.where(train_sel == True)
            train_sel = train_sel[0].tolist()
            for j, item in enumerate(train_sel):
                for _ in range(patch_num):
                    sample.append(
                        (
                            os.path.join(root, "images", imgnames[item]),
                            target[item],
                        )
                    )
        self.samples = sample
        self.transform = transform

        self.score_k = score_k
        self.bins = np.linspace(0, 100, self.score_k + 1)

        self.samples = sample
        self.transform = transform

    def _load_image(self, path, mode="RGB"):
        try:
            im = Image.open(path).convert(mode)
        except:
            print("ERROR IMG LOADED: ", path)
            random_img = np.random.rand(224, 224, 3) * 255
            im = Image.fromarray(np.uint8(random_img))
        return im

    def _depth_path_converter(self, img_path):
        directory, filename = os.path.split(img_path)
        name, ext = os.path.splitext(filename)

        depth_dir = directory.replace("images", "depths")
        depths_file = f"{name}-depth.png"
        orders_file = f"{name}-onehot_depth.npy"

        return os.path.join(depth_dir, depths_file), os.path.join(
            depth_dir, orders_file
        )

    def score_to_onehot(self, target_score):
        index = np.digitize([target_score], self.bins) - 1
        index = min(index[0], self.score_k - 1)

        one_hot = np.zeros(self.score_k, dtype=np.float32)
        one_hot[index] = 1.0
        one_hot = torch.from_numpy(one_hot).float()
        return one_hot

    def __getitem__(self, index):
        """
        Args:
            index (int): Index
        Returns:
            tuple: (sample, target) where target is class_index of the target class.
        """
        path, target = self.samples[index]
        sample = np.array(self._load_image(path)).astype(np.float32)
        cor_depth, cor_orders = (
            np.array(
                self._load_image(self._depth_path_converter(path)[0], mode="L")
            ).astype(np.float32),
            np.load(self._depth_path_converter(path)[1], allow_pickle=True).astype(
                np.float32
            ),
        )
        sample = torch.from_numpy(
            np.transpose(
                np.concatenate([sample, cor_depth[:, :, None], cor_orders], axis=-1),
                (2, 0, 1),
            )
        ).contiguous()
        sample = self.transform[0](sample)
        images, depths, orders = (
            (sample[:3, :, :] / 255.0).contiguous(),
            sample[3, :, :].contiguous(),
            sample[4:, :, :].contiguous(),
        )
        images = self.transform[2](images)
        return (
            images,
            process_inverse_depth(depths),
            orders,
            target,
            self.score_to_onehot(target),
        )

    def __len__(self):
        length = len(self.samples)
        return length


class SPAQDATASET(data.Dataset):
    def __init__(self, root, index, patch_num, transform=None, score_k=2):
        super(SPAQDATASET, self).__init__()

        self.data_path = root
        anno_folder = os.path.join(self.data_path, "Annotations")
        xlsx_file = os.path.join(anno_folder, "MOS and Image attribute scores.xlsx")
        read = pd.read_excel(xlsx_file)
        imgname = read["Image name"].values.tolist()
        mos_all = read["MOS"].values.tolist()
        for i in range(len(mos_all)):
            mos_all[i] = np.float32(mos_all[i])
        sample = []
        for _, item in enumerate(index):
            for _ in range(patch_num):
                sample.append(
                    (
                        os.path.join(
                            self.data_path,
                            # "SPAQ zip",
                            "512x384",
                            imgname[item],
                        ),
                        mos_all[item],
                    )
                )

        self.score_k = score_k
        self.bins = np.linspace(0, 100, self.score_k + 1)

        self.samples = sample
        self.transform = transform

    def _load_image(self, path, mode="RGB"):
        try:
            im = Image.open(path).convert(mode)
        except:
            print("ERROR IMG LOADED: ", path)
            random_img = np.random.rand(224, 224, 3) * 255
            im = Image.fromarray(np.uint8(random_img))
        return im

    def _depth_path_converter(self, img_path):
        directory, filename = os.path.split(img_path)
        name, ext = os.path.splitext(filename)

        depth_dir = directory.replace("512x384", "512x384_Depths")
        depths_file = f"{name}-depth.png"
        orders_file = f"{name}-onehot_depth.npy"

        return os.path.join(depth_dir, depths_file), os.path.join(
            depth_dir, orders_file
        )

    def score_to_onehot(self, target_score):
        index = np.digitize([target_score], self.bins) - 1
        index = min(index[0], self.score_k - 1)

        one_hot = np.zeros(self.score_k, dtype=np.float32)
        one_hot[index] = 1.0
        one_hot = torch.from_numpy(one_hot).float()
        return one_hot

    def __getitem__(self, index):
        """
        Args:
            index (int): Index

        Returns:
            tuple: (sample, target) where target is class_index of the target class.
        """
        path, target = self.samples[index]
        sample = np.array(self._load_image(path)).astype(np.float32)
        cor_depth, cor_orders = (
            np.array(
                self._load_image(self._depth_path_converter(path)[0], mode="L")
            ).astype(np.float32),
            np.load(self._depth_path_converter(path)[1], allow_pickle=True).astype(
                np.float32
            ),
        )
        sample = torch.from_numpy(
            np.transpose(
                np.concatenate([sample, cor_depth[:, :, None], cor_orders], axis=-1),
                (2, 0, 1),
            )
        ).contiguous()
        sample = self.transform[0](sample)
        images, depths, orders = (
            (sample[:3, :, :] / 255.0).contiguous(),
            sample[3, :, :].contiguous(),
            sample[4:, :, :].contiguous(),
        )
        images = self.transform[2](images)
        return (
            images,
            process_inverse_depth(depths),
            orders,
            target,
            self.score_to_onehot(target),
        )

    def __len__(self):
        length = len(self.samples)
        return length


class FBLIVEFolder(data.Dataset):
    def __init__(self, root, index, patch_num, transform=None, score_k=2):
        imgname = []
        mos_all = []
        csv_file = os.path.join(root, "labels_image.csv")
        with open(csv_file) as f:
            reader = csv.DictReader(f)
            for row in reader:
                imgname.append(row["name"])
                mos = np.float32(row["mos"])
                mos_all.append(mos)

        sample = []
        for i, item in enumerate(index):
            for aug in range(patch_num):
                sample.append(
                    (os.path.join(root, "database", imgname[item]), mos_all[item])
                )

        self.score_k = score_k
        self.bins = np.linspace(0, 100, self.score_k + 1)

        self.samples = sample
        self.transform = transform

    def _load_image(self, path, mode="RGB"):
        try:
            im = Image.open(path).convert(mode)
        except:
            print("ERROR IMG LOADED: ", path)
            random_img = np.random.rand(224, 224, 3) * 255
            im = Image.fromarray(np.uint8(random_img))
        return im

    def _depth_path_converter(self, img_path):
        directory, filename = os.path.split(img_path)
        name, ext = os.path.splitext(filename)

        depth_dir = directory.replace("database", "gen_depths")
        depths_file = f"{name}-depth.png"
        orders_file = f"{name}-onehot_depth.npy"

        return os.path.join(depth_dir, depths_file), os.path.join(
            depth_dir, orders_file
        )

    def score_to_onehot(self, target_score):
        index = np.digitize([target_score], self.bins) - 1
        index = min(index[0], self.score_k - 1)

        one_hot = np.zeros(self.score_k, dtype=np.float32)
        one_hot[index] = 1.0
        one_hot = torch.from_numpy(one_hot).float()
        return one_hot

    def __getitem__(self, index):
        """
        Args:
            index (int): Index
        Returns:
            tuple: (sample, target) where target is class_index of the target class.
        """
        path, target = self.samples[index]
        sample = np.array(self._load_image(path)).astype(np.float32)
        cor_depth, cor_orders = (
            np.array(
                self._load_image(self._depth_path_converter(path)[0], mode="L")
            ).astype(np.float32),
            np.load(self._depth_path_converter(path)[1], allow_pickle=True).astype(
                np.float32
            ),
        )
        if cor_depth.shape[0] != sample.shape[0]:
            cor_depth = np.transpose(cor_depth, (1, 0))
            cor_orders = np.transpose(cor_orders, (1, 0, 2))
        sample = torch.from_numpy(
            np.transpose(
                np.concatenate([sample, cor_depth[:, :, None], cor_orders], axis=-1),
                (2, 0, 1),
            )
        ).contiguous()
        sample = self.transform[0](sample)
        images, depths, orders = (
            (sample[:3, :, :] / 255.0).contiguous(),
            sample[3, :, :].contiguous(),
            sample[4:, :, :].contiguous(),
        )
        images = self.transform[2](images)
        return (
            images,
            process_inverse_depth(depths),
            orders,
            target,
            self.score_to_onehot(target),
        )

    def __len__(self):
        length = len(self.samples)
        return length


class BIDDATASET(data.Dataset):

    def __init__(self, root, index, patch_num, transform, score_k=2):

        imgname = []
        mos_all = []

        xls_file = os.path.join(root, "DatabaseGrades.xlsx")
        workbook = load_workbook(xls_file)
        booksheet = workbook.active
        rows = booksheet.rows
        count = 1
        for row in rows:
            count += 1
            img_num = booksheet.cell(row=count, column=1).value
            img_name = "DatabaseImage%04d.JPG" % (img_num)
            imgname.append(img_name)
            mos = booksheet.cell(row=count, column=2).value
            mos = np.float32(mos)
            mos_all.append(mos)
            if count == 587:
                break

        sample = []
        for i, item in enumerate(index):
            for aug in range(patch_num):
                sample.append((os.path.join(root, imgname[item]), mos_all[item]))

        self.score_k = score_k
        self.bins = np.linspace(0, 100, self.score_k + 1)

        self.samples = sample
        self.transform = transform

    def _load_image(self, path, mode="RGB"):
        try:
            im = Image.open(path).convert(mode)
        except:
            print("ERROR IMG LOADED: ", path)
            random_img = np.random.rand(224, 224, 3) * 255
            im = Image.fromarray(np.uint8(random_img))
        return im

    def _depth_path_converter(self, img_path):
        directory, filename = os.path.split(img_path)
        name, ext = os.path.splitext(filename)

        depth_dir = directory.replace("ImageDatabase", "ImageDatabase_depth")
        depths_file = f"{name}-depth.png"
        orders_file = f"{name}-onehot_depth.npy"

        return os.path.join(depth_dir, depths_file), os.path.join(
            depth_dir, orders_file
        )

    def score_to_onehot(self, target_score):
        index = np.digitize([target_score], self.bins) - 1
        index = min(index[0], self.score_k - 1)

        one_hot = np.zeros(self.score_k, dtype=np.float32)
        one_hot[index] = 1.0
        one_hot = torch.from_numpy(one_hot).float()
        return one_hot

    def __getitem__(self, index):
        """
        Args:
            index (int): Index

        Returns:
            tuple: (sample, target) where target is class_index of the target class.
        """
        path, target = self.samples[index]
        sample = np.array(self._load_image(path)).astype(np.float32)
        cor_depth, cor_orders = (
            np.array(
                self._load_image(self._depth_path_converter(path)[0], mode="L")
            ).astype(np.float32),
            np.load(self._depth_path_converter(path)[1], allow_pickle=True).astype(
                np.float32
            ),
        )
        sample = torch.from_numpy(
            np.transpose(
                np.concatenate([sample, cor_depth[:, :, None], cor_orders], axis=-1),
                (2, 0, 1),
            )
        ).contiguous()
        sample = self.transform[0](sample)
        images, depths, orders = (
            (sample[:3, :, :] / 255.0).contiguous(),
            sample[3, :, :].contiguous(),
            sample[4:, :, :].contiguous(),
        )
        images = self.transform[2](images)
        return (
            images,
            process_inverse_depth(depths),
            orders,
            target,
            self.score_to_onehot(target),
        )

    def __len__(self):
        length = len(self.samples)
        return length


def process_inverse_depth(depth):
    depth = (255 - depth) / 255.0

    # min_val = depth.min()
    # max_val = depth.max()
    # depth = (depth - min_val) / (max_val - min_val)
    # depth = 1.0 - depth

    mean = 0.485
    std = 0.229

    depth = (depth - mean) / std

    return depth
